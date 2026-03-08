"""
Revenue Analytics Dashboard - Comprehensive Earnings API
Admin-only endpoints for tracking all revenue, subscriptions, and top-ups
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from datetime import datetime, timezone, timedelta
from typing import Optional, List
import uuid
import logging
import csv
import io
from fastapi.responses import StreamingResponse, Response

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from shared import db, get_admin_user, logger

router = APIRouter(prefix="/revenue-analytics", tags=["Revenue Analytics"])


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_date(date_str: Optional[str]) -> Optional[datetime]:
    """Parse date string to datetime"""
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
    except Exception:
        try:
            return datetime.strptime(date_str, "%Y-%m-%d")
        except Exception:
            return None


async def get_user_location(user_id: str) -> dict:
    """Get user's last known location from login activity"""
    login = await db.login_activity.find_one(
        {"user_id": user_id, "location": {"$exists": True}},
        {"_id": 0, "location": 1, "ip_address": 1, "device_type": 1},
        sort=[("timestamp", -1)]
    )
    if login:
        return {
            "location": login.get("location", {}),
            "ip_address": login.get("ip_address"),
            "device_type": login.get("device_type")
        }
    return {"location": {}, "ip_address": None, "device_type": None}


def get_payment_status_description(order: dict, user: dict = None) -> dict:
    """
    Generate user-friendly description for payment status
    Returns dict with reason, description, and action_needed
    """
    status = order.get("status", "").upper()
    product_id = order.get("productId", "")
    product_name = order.get("productName", product_id)
    user_name = user.get("name", "Unknown User") if user else "Unknown User"
    user_email = order.get("userEmail") or (user.get("email") if user else "Unknown")
    amount = order.get("amount", 0) / 100
    created_at = order.get("createdAt", "")
    failure_reason = order.get("failureReason", "")
    gateway = order.get("gateway", "cashfree")
    
    # Calculate time since created
    time_info = ""
    try:
        created_dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        time_diff = datetime.now(timezone.utc) - created_dt
        if time_diff.days > 0:
            time_info = f"{time_diff.days} days ago"
        elif time_diff.seconds > 3600:
            time_info = f"{time_diff.seconds // 3600} hours ago"
        else:
            time_info = f"{time_diff.seconds // 60} minutes ago"
    except:
        time_info = "recently"
    
    # Determine product type
    if product_id in ["weekly", "monthly", "quarterly", "yearly"]:
        product_type = "subscription"
        product_display = f"{product_id.capitalize()} Subscription"
    elif product_id in ["starter", "creator", "pro"]:
        product_type = "topup"
        product_display = f"{product_id.capitalize()} Credit Pack"
    else:
        product_type = "other"
        product_display = product_name or "Unknown Product"
    
    result = {
        "status": status,
        "reason": "",
        "description": "",
        "action_needed": "",
        "user_context": f"{user_name} ({user_email})",
        "product_context": product_display,
        "amount_context": f"₹{amount:,.2f}",
        "time_context": time_info,
        "severity": "info"
    }
    
    if status == "PENDING":
        # Analyze why payment is pending
        if time_info and "days" in time_info:
            days = int(time_info.split()[0])
            if days > 7:
                result["reason"] = "Abandoned Checkout"
                result["description"] = f"{user_name} started buying {product_display} ({time_info}) but never completed payment. They may have had second thoughts or faced payment issues."
                result["action_needed"] = "Consider sending a reminder email or offering a discount to recover this sale."
                result["severity"] = "warning"
            elif days > 1:
                result["reason"] = "Payment Not Completed"
                result["description"] = f"{user_name} initiated {product_display} purchase {time_info} but payment is still pending. User may have closed browser or had payment gateway issues."
                result["action_needed"] = "Check if user needs assistance. They may have faced payment errors."
                result["severity"] = "warning"
            else:
                result["reason"] = "Awaiting Payment"
                result["description"] = f"{user_name} started buying {product_display} {time_info}. Payment is in progress through {gateway.capitalize()}."
                result["action_needed"] = "Wait for payment gateway callback. If still pending after 24 hours, investigate."
                result["severity"] = "info"
        else:
            result["reason"] = "Payment Initiated"
            result["description"] = f"{user_name} just started purchasing {product_display}. Waiting for payment confirmation from {gateway.capitalize()}."
            result["action_needed"] = "No action needed yet. Payment should complete shortly."
            result["severity"] = "info"
    
    elif status == "FAILED":
        result["severity"] = "error"
        if failure_reason:
            result["reason"] = "Payment Failed"
            result["description"] = f"{user_name} tried to buy {product_display} but payment failed. Gateway error: {failure_reason}"
            result["action_needed"] = "Reach out to user and help them retry. Check if their card/UPI has issues."
        else:
            result["reason"] = "Payment Declined"
            result["description"] = f"{user_name}'s payment for {product_display} was declined by the payment gateway."
            result["action_needed"] = "Contact user to try again with a different payment method."
    
    elif status == "CANCELLED":
        result["severity"] = "warning"
        result["reason"] = "User Cancelled"
        result["description"] = f"{user_name} cancelled the payment for {product_display} during checkout."
        result["action_needed"] = "Follow up with user to understand why they cancelled. May need support."
    
    elif status == "REFUNDED":
        result["severity"] = "info"
        refund_amount = order.get("refundAmount", amount)
        result["reason"] = "Payment Refunded"
        result["description"] = f"₹{refund_amount:,.2f} was refunded to {user_name} for {product_display}."
        result["action_needed"] = "No action needed. Refund processed successfully."
    
    elif status == "PARTIALLY_REFUNDED":
        result["severity"] = "info"
        refund_amount = order.get("refundAmount", 0)
        result["reason"] = "Partial Refund"
        result["description"] = f"₹{refund_amount:,.2f} was partially refunded to {user_name} for {product_display}."
        result["action_needed"] = "Verify if full refund was needed or partial was correct."
    
    elif status == "PAID":
        result["severity"] = "success"
        result["reason"] = "Payment Successful"
        result["description"] = f"{user_name} successfully purchased {product_display}."
        result["action_needed"] = "None - payment completed successfully!"
    
    else:
        result["reason"] = f"Status: {status}"
        result["description"] = f"Order from {user_name} for {product_display} has status: {status}"
        result["action_needed"] = "Investigate this status in payment gateway dashboard."
    
    return result


# =============================================================================
# SUMMARY DASHBOARD ENDPOINT
# =============================================================================

@router.get("/summary")
async def get_revenue_summary(
    admin: dict = Depends(get_admin_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Get comprehensive revenue summary with all key metrics
    """
    try:
        # Date filtering
        end_dt = parse_date(end_date) or datetime.now(timezone.utc)
        start_dt = parse_date(start_date) or (end_dt - timedelta(days=30))
        
        start_iso = start_dt.isoformat()
        end_iso = end_dt.isoformat()
        
        date_filter = {
            "createdAt": {"$gte": start_iso, "$lte": end_iso}
        }
        
        # Total revenue (all time)
        all_time_pipeline = [
            {"$match": {"status": "PAID"}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}
        ]
        all_time_result = await db.orders.aggregate(all_time_pipeline).to_list(1)
        total_revenue_all_time = (all_time_result[0]["total"] / 100) if all_time_result else 0
        total_orders_all_time = all_time_result[0]["count"] if all_time_result else 0
        
        # Revenue in period
        period_pipeline = [
            {"$match": {"status": "PAID", **date_filter}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}
        ]
        period_result = await db.orders.aggregate(period_pipeline).to_list(1)
        period_revenue = (period_result[0]["total"] / 100) if period_result else 0
        period_orders = period_result[0]["count"] if period_result else 0
        
        # Subscription revenue
        subscription_types = ["weekly", "monthly", "quarterly", "yearly"]
        sub_pipeline = [
            {"$match": {"status": "PAID", "productId": {"$in": subscription_types}}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}
        ]
        sub_result = await db.orders.aggregate(sub_pipeline).to_list(1)
        subscription_revenue = (sub_result[0]["total"] / 100) if sub_result else 0
        subscription_count = sub_result[0]["count"] if sub_result else 0
        
        # Top-up revenue (credit packs)
        topup_types = ["starter", "creator", "pro"]
        topup_pipeline = [
            {"$match": {"status": "PAID", "productId": {"$in": topup_types}}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}
        ]
        topup_result = await db.orders.aggregate(topup_pipeline).to_list(1)
        topup_revenue = (topup_result[0]["total"] / 100) if topup_result else 0
        topup_count = topup_result[0]["count"] if topup_result else 0
        
        # Pending payments
        pending_count = await db.orders.count_documents({"status": "PENDING"})
        pending_pipeline = [
            {"$match": {"status": "PENDING"}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
        ]
        pending_result = await db.orders.aggregate(pending_pipeline).to_list(1)
        pending_amount = (pending_result[0]["total"] / 100) if pending_result else 0
        
        # Failed payments
        failed_count = await db.orders.count_documents({"status": {"$in": ["FAILED", "CANCELLED"]}})
        failed_pipeline = [
            {"$match": {"status": {"$in": ["FAILED", "CANCELLED"]}}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
        ]
        failed_result = await db.orders.aggregate(failed_pipeline).to_list(1)
        failed_amount = (failed_result[0]["total"] / 100) if failed_result else 0
        
        # Refunded payments
        refunded_pipeline = [
            {"$match": {"status": {"$in": ["REFUNDED", "PARTIALLY_REFUNDED"]}}},
            {"$group": {"_id": None, "total": {"$sum": "$refundAmount"}, "count": {"$sum": 1}}}
        ]
        refunded_result = await db.orders.aggregate(refunded_pipeline).to_list(1)
        refunded_amount = refunded_result[0]["total"] if refunded_result else 0
        refunded_count = refunded_result[0]["count"] if refunded_result else 0
        
        # Active subscribers (users with active subscriptions)
        # Check users who have subscription in the last period based on their plan
        active_subscribers = await db.users.count_documents({
            "plan": {"$in": subscription_types}
        })
        
        # Expired subscribers (rough estimate based on last payment date)
        # This is users who had subscriptions but haven't paid in a while
        thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
        expired_query = {
            "status": "PAID",
            "productId": {"$in": subscription_types},
            "paidAt": {"$lt": thirty_days_ago}
        }
        expired_subscribers = await db.orders.distinct("userId", expired_query)
        expired_count = len(expired_subscribers)
        
        # Net revenue (total - refunded)
        net_revenue = total_revenue_all_time - refunded_amount
        
        return {
            "success": True,
            "summary": {
                "totalRevenueAllTime": round(total_revenue_all_time, 2),
                "totalOrdersAllTime": total_orders_all_time,
                "periodRevenue": round(period_revenue, 2),
                "periodOrders": period_orders,
                "subscriptionRevenue": round(subscription_revenue, 2),
                "subscriptionCount": subscription_count,
                "topupRevenue": round(topup_revenue, 2),
                "topupCount": topup_count,
                "netRevenue": round(net_revenue, 2),
                "pendingPayments": {
                    "count": pending_count,
                    "amount": round(pending_amount, 2)
                },
                "failedPayments": {
                    "count": failed_count,
                    "amount": round(failed_amount, 2)
                },
                "refundedPayments": {
                    "count": refunded_count,
                    "amount": round(refunded_amount, 2)
                },
                "activeSubscribers": active_subscribers,
                "expiredSubscribers": expired_count
            },
            "period": {
                "start": start_iso,
                "end": end_iso
            }
        }
    except Exception as e:
        logger.error(f"Revenue summary error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# PENDING PAYMENTS ANALYSIS
# =============================================================================

@router.get("/pending-analysis")
async def get_pending_payments_analysis(
    admin: dict = Depends(get_admin_user)
):
    """
    Get detailed analysis of all pending payments with user-friendly explanations
    """
    try:
        # Get all pending orders
        pending_orders = await db.orders.find(
            {"status": "PENDING"},
            {"_id": 0}
        ).sort("createdAt", -1).to_list(100)
        
        # Categorize pending payments
        abandoned = []  # > 7 days old
        stale = []  # 1-7 days old
        recent = []  # < 1 day old
        
        total_amount = 0
        
        for order in pending_orders:
            user_id = order.get("userId")
            user = await db.users.find_one({"id": user_id}, {"_id": 0, "name": 1, "email": 1})
            
            status_info = get_payment_status_description(order, user)
            
            # Calculate age
            try:
                created_dt = datetime.fromisoformat(order.get("createdAt", "").replace('Z', '+00:00'))
                age_days = (datetime.now(timezone.utc) - created_dt).days
            except:
                age_days = 0
            
            amount = order.get("amount", 0) / 100
            total_amount += amount
            
            payment_info = {
                "orderId": order.get("order_id", order.get("id")),
                "userId": user_id,
                "userName": user.get("name") if user else "Unknown",
                "userEmail": order.get("userEmail") or (user.get("email") if user else "Unknown"),
                "productId": order.get("productId"),
                "productName": order.get("productName"),
                "amount": amount,
                "currency": order.get("currency", "INR"),
                "createdAt": order.get("createdAt"),
                "ageDays": age_days,
                "statusInfo": status_info
            }
            
            if age_days > 7:
                abandoned.append(payment_info)
            elif age_days >= 1:
                stale.append(payment_info)
            else:
                recent.append(payment_info)
        
        return {
            "success": True,
            "analysis": {
                "totalPending": len(pending_orders),
                "totalAmount": round(total_amount, 2),
                "categories": {
                    "abandoned": {
                        "count": len(abandoned),
                        "amount": round(sum(p["amount"] for p in abandoned), 2),
                        "description": "Checkouts abandoned over 7 days ago. Users likely had second thoughts or faced issues.",
                        "action": "Send reminder emails with discount codes to recover these sales.",
                        "payments": abandoned
                    },
                    "stale": {
                        "count": len(stale),
                        "amount": round(sum(p["amount"] for p in stale), 2),
                        "description": "Payments started 1-7 days ago but never completed.",
                        "action": "Reach out to users to see if they need help completing payment.",
                        "payments": stale
                    },
                    "recent": {
                        "count": len(recent),
                        "amount": round(sum(p["amount"] for p in recent), 2),
                        "description": "Recent checkouts in progress. Payment may complete soon.",
                        "action": "No immediate action needed. Wait for payment gateway callback.",
                        "payments": recent
                    }
                }
            }
        }
    except Exception as e:
        logger.error(f"Pending analysis error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# SUBSCRIPTION BREAKDOWN
# =============================================================================

@router.get("/subscriptions")
async def get_subscription_breakdown(
    admin: dict = Depends(get_admin_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Get detailed breakdown of subscription revenue by type
    """
    try:
        date_filter = {}
        if start_date or end_date:
            end_dt = parse_date(end_date) or datetime.now(timezone.utc)
            start_dt = parse_date(start_date) or (end_dt - timedelta(days=365))
            date_filter = {"paidAt": {"$gte": start_dt.isoformat(), "$lte": end_dt.isoformat()}}
        
        subscription_types = ["weekly", "monthly", "quarterly", "yearly"]
        
        breakdown = []
        for sub_type in subscription_types:
            pipeline = [
                {"$match": {"status": "PAID", "productId": sub_type, **date_filter}},
                {"$group": {
                    "_id": None,
                    "total": {"$sum": "$amount"},
                    "count": {"$sum": 1},
                    "uniqueUsers": {"$addToSet": "$userId"}
                }}
            ]
            result = await db.orders.aggregate(pipeline).to_list(1)
            
            if result:
                breakdown.append({
                    "type": sub_type,
                    "revenue": round(result[0]["total"] / 100, 2),
                    "count": result[0]["count"],
                    "uniqueUsers": len(result[0]["uniqueUsers"])
                })
            else:
                breakdown.append({
                    "type": sub_type,
                    "revenue": 0,
                    "count": 0,
                    "uniqueUsers": 0
                })
        
        return {
            "success": True,
            "subscriptions": breakdown,
            "total": sum(s["revenue"] for s in breakdown)
        }
    except Exception as e:
        logger.error(f"Subscription breakdown error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# TOP-UP BREAKDOWN
# =============================================================================

@router.get("/topups")
async def get_topup_breakdown(
    admin: dict = Depends(get_admin_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Get detailed breakdown of top-up/credit pack purchases
    """
    try:
        date_filter = {}
        if start_date or end_date:
            end_dt = parse_date(end_date) or datetime.now(timezone.utc)
            start_dt = parse_date(start_date) or (end_dt - timedelta(days=365))
            date_filter = {"paidAt": {"$gte": start_dt.isoformat(), "$lte": end_dt.isoformat()}}
        
        topup_types = ["starter", "creator", "pro"]
        
        breakdown = []
        for topup_type in topup_types:
            pipeline = [
                {"$match": {"status": "PAID", "productId": topup_type, **date_filter}},
                {"$group": {
                    "_id": None,
                    "total": {"$sum": "$amount"},
                    "count": {"$sum": 1},
                    "uniqueUsers": {"$addToSet": "$userId"}
                }}
            ]
            result = await db.orders.aggregate(pipeline).to_list(1)
            
            if result:
                breakdown.append({
                    "type": topup_type,
                    "revenue": round(result[0]["total"] / 100, 2),
                    "count": result[0]["count"],
                    "uniqueUsers": len(result[0]["uniqueUsers"])
                })
            else:
                breakdown.append({
                    "type": topup_type,
                    "revenue": 0,
                    "count": 0,
                    "uniqueUsers": 0
                })
        
        return {
            "success": True,
            "topups": breakdown,
            "total": sum(t["revenue"] for t in breakdown)
        }
    except Exception as e:
        logger.error(f"Top-up breakdown error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# TIME-BASED REVENUE ANALYTICS
# =============================================================================

@router.get("/trends")
async def get_revenue_trends(
    admin: dict = Depends(get_admin_user),
    period: str = Query(default="day", regex="^(day|week|month|year)$"),
    limit: int = Query(default=30, ge=1, le=365)
):
    """
    Get revenue trends by day/week/month/year
    """
    try:
        now = datetime.now(timezone.utc)
        
        if period == "day":
            date_format = "%Y-%m-%d"
            delta = timedelta(days=1)
        elif period == "week":
            date_format = "%Y-W%W"
            delta = timedelta(weeks=1)
        elif period == "month":
            date_format = "%Y-%m"
            delta = timedelta(days=30)
        else:  # year
            date_format = "%Y"
            delta = timedelta(days=365)
        
        trends = []
        for i in range(limit):
            period_start = now - (delta * (i + 1))
            period_end = now - (delta * i)
            
            pipeline = [
                {"$match": {
                    "status": "PAID",
                    "paidAt": {
                        "$gte": period_start.isoformat(),
                        "$lt": period_end.isoformat()
                    }
                }},
                {"$group": {
                    "_id": None,
                    "revenue": {"$sum": "$amount"},
                    "orders": {"$sum": 1},
                    "subscriptions": {"$sum": {"$cond": [{"$in": ["$productId", ["weekly", "monthly", "quarterly", "yearly"]]}, 1, 0]}},
                    "topups": {"$sum": {"$cond": [{"$in": ["$productId", ["starter", "creator", "pro"]]}, 1, 0]}}
                }}
            ]
            
            result = await db.orders.aggregate(pipeline).to_list(1)
            
            trends.append({
                "period": period_end.strftime(date_format),
                "periodStart": period_start.isoformat(),
                "periodEnd": period_end.isoformat(),
                "revenue": round(result[0]["revenue"] / 100, 2) if result else 0,
                "orders": result[0]["orders"] if result else 0,
                "subscriptions": result[0]["subscriptions"] if result else 0,
                "topups": result[0]["topups"] if result else 0
            })
        
        # Reverse to show oldest first
        trends.reverse()
        
        return {
            "success": True,
            "period": period,
            "trends": trends
        }
    except Exception as e:
        logger.error(f"Revenue trends error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# DETAILED TRANSACTIONS LIST
# =============================================================================

@router.get("/transactions")
async def get_transactions(
    admin: dict = Depends(get_admin_user),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=20, ge=1, le=100),
    status: Optional[str] = None,
    product_type: Optional[str] = None,
    user_email: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    country: Optional[str] = None
):
    """
    Get detailed list of all transactions with filters
    """
    try:
        # Build filter
        filter_query = {}
        
        if status:
            filter_query["status"] = status.upper()
        
        if product_type:
            if product_type == "subscription":
                filter_query["productId"] = {"$in": ["weekly", "monthly", "quarterly", "yearly"]}
            elif product_type == "topup":
                filter_query["productId"] = {"$in": ["starter", "creator", "pro"]}
            else:
                filter_query["productId"] = product_type
        
        if user_email:
            filter_query["userEmail"] = {"$regex": user_email, "$options": "i"}
        
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date
            if end_date:
                date_filter["$lte"] = end_date
            filter_query["createdAt"] = date_filter
        
        # Get total count
        total = await db.orders.count_documents(filter_query)
        
        # Get transactions
        skip = (page - 1) * limit
        orders = await db.orders.find(
            filter_query,
            {"_id": 0}
        ).sort("createdAt", -1).skip(skip).limit(limit).to_list(limit)
        
        # Enrich with user data and location
        enriched_transactions = []
        for order in orders:
            user_id = order.get("userId")
            user = await db.users.find_one({"id": user_id}, {"_id": 0, "name": 1, "email": 1, "plan": 1})
            location_data = await get_user_location(user_id)
            
            # Determine transaction type
            product_id = order.get("productId", "")
            if product_id in ["weekly", "monthly", "quarterly", "yearly"]:
                tx_type = "subscription"
            elif product_id in ["starter", "creator", "pro"]:
                tx_type = "topup"
            else:
                tx_type = "other"
            
            # Check if renewal
            previous_order = await db.orders.find_one({
                "userId": user_id,
                "status": "PAID",
                "createdAt": {"$lt": order.get("createdAt")}
            })
            is_renewal = previous_order is not None
            
            # Get user-friendly status description
            status_info = get_payment_status_description(order, user)
            
            enriched_transactions.append({
                "orderId": order.get("order_id", order.get("id")),
                "cfOrderId": order.get("cf_order_id"),
                "userId": user_id,
                "userName": user.get("name") if user else "Unknown",
                "userEmail": order.get("userEmail") or (user.get("email") if user else "Unknown"),
                "productId": product_id,
                "productName": order.get("productName"),
                "transactionType": tx_type,
                "isRenewal": is_renewal,
                "amount": round(order.get("amount", 0) / 100, 2),
                "currency": order.get("currency", "INR"),
                "credits": order.get("credits", 0),
                "status": order.get("status"),
                "statusInfo": status_info,  # NEW: User-friendly status description
                "gateway": order.get("gateway", "cashfree"),
                "createdAt": order.get("createdAt"),
                "paidAt": order.get("paidAt"),
                "refundedAt": order.get("refundedAt"),
                "refundAmount": order.get("refundAmount"),
                "failureReason": order.get("failureReason"),
                "location": location_data.get("location", {}),
                "ipAddress": location_data.get("ip_address"),
                "deviceType": location_data.get("device_type")
            })
        
        # Apply country filter if specified (post-query filter since location is enriched)
        if country:
            enriched_transactions = [
                t for t in enriched_transactions 
                if t.get("location", {}).get("country", "").lower() == country.lower()
            ]
        
        return {
            "success": True,
            "transactions": enriched_transactions,
            "pagination": {
                "page": page,
                "limit": limit,
                "total": total,
                "totalPages": (total + limit - 1) // limit
            }
        }
    except Exception as e:
        logger.error(f"Transactions list error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# USER PAYMENT HISTORY
# =============================================================================

@router.get("/user/{user_id}/history")
async def get_user_payment_history(
    user_id: str,
    admin: dict = Depends(get_admin_user)
):
    """
    Get complete payment history for a specific user
    """
    try:
        # Get user info
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Get all orders for this user
        orders = await db.orders.find(
            {"userId": user_id},
            {"_id": 0}
        ).sort("createdAt", -1).to_list(1000)
        
        # Get credit ledger
        credits = await db.credit_ledger.find(
            {"userId": user_id},
            {"_id": 0}
        ).sort("timestamp", -1).to_list(100)
        
        # Calculate totals
        total_spent = sum(o.get("amount", 0) for o in orders if o.get("status") == "PAID") / 100
        total_refunded = sum(o.get("refundAmount", 0) for o in orders if o.get("status") in ["REFUNDED", "PARTIALLY_REFUNDED"])
        subscription_orders = [o for o in orders if o.get("productId") in ["weekly", "monthly", "quarterly", "yearly"]]
        topup_orders = [o for o in orders if o.get("productId") in ["starter", "creator", "pro"]]
        
        # Get location data
        location_data = await get_user_location(user_id)
        
        return {
            "success": True,
            "user": {
                "id": user_id,
                "name": user.get("name"),
                "email": user.get("email"),
                "plan": user.get("plan"),
                "credits": user.get("credits"),
                "createdAt": user.get("createdAt"),
                "location": location_data.get("location"),
                "deviceType": location_data.get("device_type")
            },
            "summary": {
                "totalSpent": round(total_spent, 2),
                "totalRefunded": round(total_refunded, 2),
                "netSpent": round(total_spent - total_refunded, 2),
                "totalOrders": len(orders),
                "successfulOrders": len([o for o in orders if o.get("status") == "PAID"]),
                "subscriptionOrders": len(subscription_orders),
                "topupOrders": len(topup_orders)
            },
            "orders": orders,
            "creditLedger": credits[:50]  # Last 50 credit transactions
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"User payment history error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# TRANSACTION DETAIL
# =============================================================================

@router.get("/transaction/{order_id}")
async def get_transaction_detail(
    order_id: str,
    admin: dict = Depends(get_admin_user)
):
    """
    Get full details of a specific transaction
    """
    try:
        # Find order
        order = await db.orders.find_one(
            {"$or": [{"order_id": order_id}, {"id": order_id}]},
            {"_id": 0}
        )
        
        if not order:
            raise HTTPException(status_code=404, detail="Transaction not found")
        
        user_id = order.get("userId")
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        location_data = await get_user_location(user_id)
        
        # Get related refund logs
        refund_logs = await db.refund_logs.find(
            {"orderId": order_id},
            {"_id": 0}
        ).to_list(10)
        
        # Get webhook logs for this order
        webhook_logs = await db.webhook_logs.find(
            {"$or": [
                {"payload.data.order.order_id": order_id},
                {"order_id": order_id}
            ]},
            {"_id": 0}
        ).to_list(10)
        
        # Get credit ledger entry for this order
        credit_entry = await db.credit_ledger.find_one(
            {"orderId": order_id},
            {"_id": 0}
        )
        
        return {
            "success": True,
            "transaction": {
                **order,
                "amount": round(order.get("amount", 0) / 100, 2)
            },
            "user": user,
            "location": location_data,
            "refundLogs": refund_logs,
            "webhookLogs": webhook_logs,
            "creditEntry": credit_entry
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Transaction detail error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# TOP PAYING USERS
# =============================================================================

@router.get("/top-users")
async def get_top_paying_users(
    admin: dict = Depends(get_admin_user),
    limit: int = Query(default=10, ge=1, le=50)
):
    """
    Get top paying users by total revenue
    """
    try:
        pipeline = [
            {"$match": {"status": "PAID"}},
            {"$group": {
                "_id": "$userId",
                "totalSpent": {"$sum": "$amount"},
                "orderCount": {"$sum": 1},
                "lastOrder": {"$max": "$paidAt"}
            }},
            {"$sort": {"totalSpent": -1}},
            {"$limit": limit}
        ]
        
        top_users = await db.orders.aggregate(pipeline).to_list(limit)
        
        # Enrich with user data
        enriched = []
        for u in top_users:
            user = await db.users.find_one({"id": u["_id"]}, {"_id": 0, "name": 1, "email": 1, "plan": 1, "credits": 1})
            location_data = await get_user_location(u["_id"])
            
            enriched.append({
                "userId": u["_id"],
                "name": user.get("name") if user else "Unknown",
                "email": user.get("email") if user else "Unknown",
                "plan": user.get("plan") if user else "Unknown",
                "currentCredits": user.get("credits", 0) if user else 0,
                "totalSpent": round(u["totalSpent"] / 100, 2),
                "orderCount": u["orderCount"],
                "lastOrderDate": u["lastOrder"],
                "location": location_data.get("location", {})
            })
        
        return {
            "success": True,
            "topUsers": enriched
        }
    except Exception as e:
        logger.error(f"Top users error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# COUNTRY/LOCATION BREAKDOWN
# =============================================================================

@router.get("/by-location")
async def get_revenue_by_location(
    admin: dict = Depends(get_admin_user)
):
    """
    Get revenue breakdown by country/location
    """
    try:
        # Get all paid orders with user IDs
        paid_orders = await db.orders.find(
            {"status": "PAID"},
            {"_id": 0, "userId": 1, "amount": 1}
        ).to_list(10000)
        
        # Group by country
        country_revenue = {}
        for order in paid_orders:
            location_data = await get_user_location(order.get("userId", ""))
            country = location_data.get("location", {}).get("country", "Unknown")
            
            if country not in country_revenue:
                country_revenue[country] = {"revenue": 0, "orders": 0}
            
            country_revenue[country]["revenue"] += order.get("amount", 0) / 100
            country_revenue[country]["orders"] += 1
        
        # Convert to list and sort
        breakdown = [
            {"country": k, "revenue": round(v["revenue"], 2), "orders": v["orders"]}
            for k, v in country_revenue.items()
        ]
        breakdown.sort(key=lambda x: x["revenue"], reverse=True)
        
        return {
            "success": True,
            "locationBreakdown": breakdown[:20]  # Top 20 countries
        }
    except Exception as e:
        logger.error(f"Location breakdown error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# MOST PURCHASED PRODUCTS
# =============================================================================

@router.get("/popular-products")
async def get_popular_products(
    admin: dict = Depends(get_admin_user)
):
    """
    Get most purchased subscription plans and top-up packs
    """
    try:
        pipeline = [
            {"$match": {"status": "PAID"}},
            {"$group": {
                "_id": "$productId",
                "revenue": {"$sum": "$amount"},
                "count": {"$sum": 1},
                "uniqueUsers": {"$addToSet": "$userId"}
            }},
            {"$sort": {"count": -1}}
        ]
        
        products = await db.orders.aggregate(pipeline).to_list(20)
        
        result = []
        for p in products:
            product_id = p["_id"]
            if product_id in ["weekly", "monthly", "quarterly", "yearly"]:
                category = "subscription"
            elif product_id in ["starter", "creator", "pro"]:
                category = "topup"
            else:
                category = "other"
            
            result.append({
                "productId": product_id,
                "category": category,
                "revenue": round(p["revenue"] / 100, 2),
                "purchaseCount": p["count"],
                "uniqueUsers": len(p["uniqueUsers"])
            })
        
        return {
            "success": True,
            "popularProducts": result
        }
    except Exception as e:
        logger.error(f"Popular products error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# EXPORT ENDPOINTS
# =============================================================================

@router.get("/export/csv")
async def export_transactions_csv(
    admin: dict = Depends(get_admin_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None
):
    """
    Export transactions to CSV
    """
    try:
        # Build filter
        filter_query = {}
        if status:
            filter_query["status"] = status.upper()
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date
            if end_date:
                date_filter["$lte"] = end_date
            filter_query["createdAt"] = date_filter
        
        orders = await db.orders.find(filter_query, {"_id": 0}).sort("createdAt", -1).to_list(10000)
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Order ID", "User Email", "Product", "Amount (INR)", "Credits",
            "Status", "Gateway", "Created At", "Paid At", "Failure Reason"
        ])
        
        # Data
        for order in orders:
            writer.writerow([
                order.get("order_id", order.get("id")),
                order.get("userEmail", ""),
                order.get("productName", order.get("productId", "")),
                round(order.get("amount", 0) / 100, 2),
                order.get("credits", 0),
                order.get("status", ""),
                order.get("gateway", ""),
                order.get("createdAt", ""),
                order.get("paidAt", ""),
                order.get("failureReason", "")
            ])
        
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )
    except Exception as e:
        logger.error(f"CSV export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/excel")
async def export_transactions_excel(
    admin: dict = Depends(get_admin_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """
    Export transactions to Excel format (XLSX)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Build filter
        filter_query = {}
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date
            if end_date:
                date_filter["$lte"] = end_date
            filter_query["createdAt"] = date_filter
        
        orders = await db.orders.find(filter_query, {"_id": 0}).sort("createdAt", -1).to_list(10000)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        
        headers = [
            "Order ID", "User Email", "User Name", "Product", "Type",
            "Amount (INR)", "Credits", "Status", "Gateway", 
            "Created At", "Paid At", "Failure Reason"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_num, order in enumerate(orders, 2):
            user = await db.users.find_one({"id": order.get("userId")}, {"_id": 0, "name": 1})
            product_id = order.get("productId", "")
            tx_type = "Subscription" if product_id in ["weekly", "monthly", "quarterly", "yearly"] else "Top-up" if product_id in ["starter", "creator", "pro"] else "Other"
            
            ws.cell(row=row_num, column=1, value=order.get("order_id", order.get("id")))
            ws.cell(row=row_num, column=2, value=order.get("userEmail", ""))
            ws.cell(row=row_num, column=3, value=user.get("name") if user else "Unknown")
            ws.cell(row=row_num, column=4, value=order.get("productName", product_id))
            ws.cell(row=row_num, column=5, value=tx_type)
            ws.cell(row=row_num, column=6, value=round(order.get("amount", 0) / 100, 2))
            ws.cell(row=row_num, column=7, value=order.get("credits", 0))
            ws.cell(row=row_num, column=8, value=order.get("status", ""))
            ws.cell(row=row_num, column=9, value=order.get("gateway", ""))
            ws.cell(row=row_num, column=10, value=order.get("createdAt", ""))
            ws.cell(row=row_num, column=11, value=order.get("paidAt", ""))
            ws.cell(row=row_num, column=12, value=order.get("failureReason", ""))
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed for Excel export")
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
